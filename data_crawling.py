import os                           
import pandas as pd                 
import openpyxl

from openpyxl import load_workbook


# Finding path for raw data directory
path = os.getcwd()      
files = os.listdir(path+"/raw_data")
file = "disease.xlsx"


print("************************LOADING WORKBOOK\n\n\n")

wb = load_workbook(path+r"/raw_data/"+file)
ws = wb.active

print("************************BEFORE UNMERGE\n\n\n")

ws.unmerge_cells('A1:R200')

print("************************COMPLETED UNMERGE\n\n\n")

data = pd.read_excel(path+"/raw_data/"+file,'sheet 1')
data_list = data.values.tolist()
print(data_list)

clean_data_list = []
for dl in data_list:
    clean_data_list.append([x for x in dl if str(x) != 'nan'])
print(clean_data_list)


# Convert excel to list
# for file in files:
#     print(file)
#     wb = load_workbook(path+"/raw_data"+file)

#     ws = wb.activate
#     ws.unmerge_cells('A1:R200')



#     data = pd.read_excel(path+"/raw_data/"+file,'sheet 1')
#     data_list = data.values.tolist()
#     print(data_list)

#     clean_data_list = []
#     for dl in data_list:
#         clean_data_list.append([x for x in dl if str(x) != 'nan'])  
    
    #print(clean_data_list)
    #df = df.append(data)